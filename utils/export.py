"""
Export utilities for generating Excel and PDF files from mistakes data.
"""
import io
from typing import List
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from typing import TYPE_CHECKING
if TYPE_CHECKING:
    from sqlalchemy.orm import Session
    from app.models import GREMistake, Vocabulary


def export_to_excel(mistakes: List['GREMistake']) -> bytes:
    """Export mistakes to Excel format."""
    wb = Workbook()
    ws = wb.active
    ws.title = "GRE Mistakes"
    
    # Header row
    headers = [
        "ID", "Section", "Topic", "Sub Topic", "KMF Section", "KMF Problem Set",
        "Error Type", "Problem Statement", "Solution", "What Did I Do Wrong",
        "What Will I Do Next Time", "Additional Techniques", "Relevant Concept",
        "Total Attempts", "Repetition Count", "Mastered", "Got Correct",
        "Next Review Date", "Created At"
    ]
    
    # Style header
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Data rows
    for row_num, mistake in enumerate(mistakes, 2):
        ws.cell(row=row_num, column=1, value=mistake.id).border = border
        ws.cell(row=row_num, column=2, value=mistake.section).border = border
        ws.cell(row=row_num, column=3, value=mistake.topic).border = border
        ws.cell(row=row_num, column=4, value=mistake.sub_topic or "").border = border
        ws.cell(row=row_num, column=5, value=mistake.kmf_section or "").border = border
        ws.cell(row=row_num, column=6, value=mistake.kmf_problem_set or "").border = border
        ws.cell(row=row_num, column=7, value=mistake.error_type).border = border
        ws.cell(row=row_num, column=8, value=mistake.problem_statement_text or "").border = border
        ws.cell(row=row_num, column=9, value=mistake.solution_text or "").border = border
        ws.cell(row=row_num, column=10, value=mistake.what_did_i_do_wrong or "").border = border
        ws.cell(row=row_num, column=11, value=mistake.what_will_i_do_next_time or "").border = border
        ws.cell(row=row_num, column=12, value=mistake.additional_techniques or "").border = border
        ws.cell(row=row_num, column=13, value=mistake.relevant_concept or "").border = border
        ws.cell(row=row_num, column=14, value=mistake.total_attempts or 0).border = border
        ws.cell(row=row_num, column=15, value=mistake.repetition_count or 0).border = border
        ws.cell(row=row_num, column=16, value="Yes" if mistake.mastered else "No").border = border
        ws.cell(row=row_num, column=17, value="Yes" if mistake.got_correct else "No").border = border
        ws.cell(row=row_num, column=18, value=mistake.next_review_date.strftime("%Y-%m-%d %H:%M:%S") if mistake.next_review_date else "").border = border
        ws.cell(row=row_num, column=19, value=mistake.created_at.strftime("%Y-%m-%d %H:%M:%S") if mistake.created_at else "").border = border
        
        # Set alignment for text cells
        for col in range(1, 20):
            cell = ws.cell(row=row_num, column=col)
            if col in [8, 9, 10, 11, 12, 13]:  # Text fields
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Auto-adjust column widths
    column_widths = {
        1: 8,   # ID
        2: 12,  # Section
        3: 20,  # Topic
        4: 20,  # Sub Topic
        5: 12,  # KMF Section
        6: 15,  # KMF Problem Set
        7: 15,  # Error Type
        8: 40,  # Problem Statement
        9: 40,  # Solution
        10: 30, # What Did I Do Wrong
        11: 30, # What Will I Do Next Time
        12: 30, # Additional Techniques
        13: 30, # Relevant Concept
        14: 12, # Total Attempts
        15: 15, # Repetition Count
        16: 10, # Mastered
        17: 12, # Got Correct
        18: 20, # Next Review Date
        19: 20, # Created At
    }
    
    for col_num, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def export_to_pdf(mistakes: List['GREMistake']) -> bytes:
    """Export mistakes to PDF format."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    # Container for the 'Flowable' objects
    elements = []
    
    # Define styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#366092'),
        spaceAfter=12,
        spaceBefore=20
    )
    
    normal_style = styles['Normal']
    normal_style.fontSize = 10
    normal_style.leading = 14
    
    # Title
    elements.append(Paragraph("GRE Mistakes Export", title_style))
    elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", normal_style))
    elements.append(Paragraph(f"Total Mistakes: {len(mistakes)}", normal_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Group mistakes by section
    quant_mistakes = [m for m in mistakes if m.section == "Quant"]
    verbal_mistakes = [m for m in mistakes if m.section == "Verbal"]
    
    # Quant Section
    if quant_mistakes:
        elements.append(Paragraph("Quantitative Section", heading_style))
        for idx, mistake in enumerate(quant_mistakes, 1):
            elements.append(Paragraph(f"<b>Mistake {idx} (ID: {mistake.id})</b>", normal_style))
            elements.append(Paragraph(f"<b>Topic:</b> {mistake.topic}", normal_style))
            if mistake.sub_topic:
                elements.append(Paragraph(f"<b>Sub Topic:</b> {mistake.sub_topic}", normal_style))
            if mistake.kmf_section:
                elements.append(Paragraph(f"<b>KMF Section:</b> {mistake.kmf_section}", normal_style))
            if mistake.kmf_problem_set:
                elements.append(Paragraph(f"<b>KMF Problem Set:</b> {mistake.kmf_problem_set}", normal_style))
            elements.append(Paragraph(f"<b>Error Type:</b> {mistake.error_type}", normal_style))
            
            if mistake.problem_statement_text:
                elements.append(Paragraph("<b>Problem Statement:</b>", normal_style))
                elements.append(Paragraph(mistake.problem_statement_text[:500] + ("..." if len(mistake.problem_statement_text) > 500 else ""), normal_style))
            
            if mistake.solution_text:
                elements.append(Paragraph("<b>Solution:</b>", normal_style))
                elements.append(Paragraph(mistake.solution_text[:500] + ("..." if len(mistake.solution_text) > 500 else ""), normal_style))
            
            if mistake.what_did_i_do_wrong:
                elements.append(Paragraph("<b>What Did I Do Wrong:</b>", normal_style))
                elements.append(Paragraph(mistake.what_did_i_do_wrong, normal_style))
            
            if mistake.what_will_i_do_next_time:
                elements.append(Paragraph("<b>What Will I Do Next Time:</b>", normal_style))
                elements.append(Paragraph(mistake.what_will_i_do_next_time, normal_style))
            
            elements.append(Paragraph(f"<b>Attempts:</b> {mistake.total_attempts or 0} | <b>Repetitions:</b> {mistake.repetition_count or 0}/5 | <b>Mastered:</b> {'Yes' if mistake.mastered else 'No'}", normal_style))
            elements.append(Spacer(1, 0.2*inch))
        
        if verbal_mistakes:
            elements.append(PageBreak())
    
    # Verbal Section
    if verbal_mistakes:
        elements.append(Paragraph("Verbal Section", heading_style))
        for idx, mistake in enumerate(verbal_mistakes, 1):
            elements.append(Paragraph(f"<b>Mistake {idx} (ID: {mistake.id})</b>", normal_style))
            elements.append(Paragraph(f"<b>Topic:</b> {mistake.topic}", normal_style))
            if mistake.sub_topic:
                elements.append(Paragraph(f"<b>Sub Topic:</b> {mistake.sub_topic}", normal_style))
            if mistake.kmf_section:
                elements.append(Paragraph(f"<b>KMF Section:</b> {mistake.kmf_section}", normal_style))
            if mistake.kmf_problem_set:
                elements.append(Paragraph(f"<b>KMF Problem Set:</b> {mistake.kmf_problem_set}", normal_style))
            elements.append(Paragraph(f"<b>Error Type:</b> {mistake.error_type}", normal_style))
            
            if mistake.problem_statement_text:
                elements.append(Paragraph("<b>Problem Statement:</b>", normal_style))
                elements.append(Paragraph(mistake.problem_statement_text[:500] + ("..." if len(mistake.problem_statement_text) > 500 else ""), normal_style))
            
            if mistake.solution_text:
                elements.append(Paragraph("<b>Solution:</b>", normal_style))
                elements.append(Paragraph(mistake.solution_text[:500] + ("..." if len(mistake.solution_text) > 500 else ""), normal_style))
            
            if mistake.what_did_i_do_wrong:
                elements.append(Paragraph("<b>What Did I Do Wrong:</b>", normal_style))
                elements.append(Paragraph(mistake.what_did_i_do_wrong, normal_style))
            
            if mistake.what_will_i_do_next_time:
                elements.append(Paragraph("<b>What Will I Do Next Time:</b>", normal_style))
                elements.append(Paragraph(mistake.what_will_i_do_next_time, normal_style))
            
            elements.append(Paragraph(f"<b>Attempts:</b> {mistake.total_attempts or 0} | <b>Repetitions:</b> {mistake.repetition_count or 0}/5 | <b>Mastered:</b> {'Yes' if mistake.mastered else 'No'}", normal_style))
            elements.append(Spacer(1, 0.2*inch))
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def export_vocabulary_to_excel(vocabulary: List['Vocabulary']) -> bytes:
    """Export vocabulary to Excel format."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Vocabulary"
    
    # Header row
    headers = [
        "ID", "Word", "Set No", "Category", "Meaning", "Synonym",
        "Sentences", "Genre", "Tags", "Source Mistake ID", "Created At"
    ]
    
    # Style header
    header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Data rows
    for row_num, vocab in enumerate(vocabulary, 2):
        ws.cell(row=row_num, column=1, value=vocab.id).border = border
        ws.cell(row=row_num, column=2, value=vocab.word).border = border
        ws.cell(row=row_num, column=3, value=vocab.set_no or "").border = border
        ws.cell(row=row_num, column=4, value=vocab.category or "").border = border
        ws.cell(row=row_num, column=5, value=vocab.meaning).border = border
        ws.cell(row=row_num, column=6, value=vocab.synonym or "").border = border
        ws.cell(row=row_num, column=7, value=vocab.sentences or "").border = border
        ws.cell(row=row_num, column=8, value=vocab.genre or "").border = border
        ws.cell(row=row_num, column=9, value=", ".join(vocab.tags) if vocab.tags else "").border = border
        ws.cell(row=row_num, column=10, value=vocab.source_mistake_id or "").border = border
        ws.cell(row=row_num, column=11, value=vocab.created_at.strftime("%Y-%m-%d %H:%M:%S") if vocab.created_at else "").border = border
        
        # Set alignment for text cells
        for col in range(1, 12):
            cell = ws.cell(row=row_num, column=col)
            if col in [5, 6, 7]:  # Meaning, Synonym, Sentences
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Auto-adjust column widths
    column_widths = {
        1: 8,   # ID
        2: 20,  # Word
        3: 10,  # Set No
        4: 15,  # Category
        5: 40,  # Meaning
        6: 30,  # Synonym
        7: 40,  # Sentences
        8: 15,  # Genre
        9: 25,  # Tags
        10: 15, # Source Mistake ID
        11: 20, # Created At
    }
    
    for col_num, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def export_vocabulary_to_pdf(vocabulary: List['Vocabulary']) -> bytes:
    """Export vocabulary to PDF format."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    # Container for the 'Flowable' objects
    elements = []
    
    # Define styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#7C3AED'),
        spaceAfter=12,
        spaceBefore=20
    )
    
    normal_style = styles['Normal']
    normal_style.fontSize = 10
    normal_style.leading = 14
    
    # Title
    elements.append(Paragraph("Vocabulary Export", title_style))
    elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", normal_style))
    elements.append(Paragraph(f"Total Vocabulary Entries: {len(vocabulary)}", normal_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Group vocabulary by set_no if available
    grouped_vocab = {}
    ungrouped_vocab = []
    
    for vocab in vocabulary:
        if vocab.set_no:
            if vocab.set_no not in grouped_vocab:
                grouped_vocab[vocab.set_no] = []
            grouped_vocab[vocab.set_no].append(vocab)
        else:
            ungrouped_vocab.append(vocab)
    
    # Grouped by Set No
    if grouped_vocab:
        for set_no in sorted(grouped_vocab.keys()):
            vocab_list = grouped_vocab[set_no]
            elements.append(Paragraph(f"Set {set_no} ({len(vocab_list)} entries)", heading_style))
            
            for idx, vocab in enumerate(vocab_list, 1):
                elements.append(Paragraph(f"<b>{idx}. {vocab.word}</b>", normal_style))
                elements.append(Paragraph(f"<b>Meaning:</b> {vocab.meaning}", normal_style))
                
                if vocab.synonym:
                    elements.append(Paragraph(f"<b>Synonym:</b> {vocab.synonym}", normal_style))
                
                if vocab.sentences:
                    elements.append(Paragraph(f"<b>Example Sentences:</b> {vocab.sentences}", normal_style))
                
                # Metadata
                metadata_parts = []
                if vocab.category:
                    metadata_parts.append(f"Category: {vocab.category}")
                if vocab.genre:
                    metadata_parts.append(f"Genre: {vocab.genre}")
                if vocab.tags:
                    metadata_parts.append(f"Tags: {', '.join(vocab.tags)}")
                
                if metadata_parts:
                    elements.append(Paragraph(f"<i>{' | '.join(metadata_parts)}</i>", normal_style))
                
                elements.append(Spacer(1, 0.15*inch))
            
            if set_no < max(grouped_vocab.keys()) or ungrouped_vocab:
                elements.append(PageBreak())
    
    # Ungrouped vocabulary
    if ungrouped_vocab:
        if grouped_vocab:
            elements.append(Paragraph("Other Vocabulary", heading_style))
        else:
            elements.append(Paragraph("Vocabulary Entries", heading_style))
        
        for idx, vocab in enumerate(ungrouped_vocab, 1):
            elements.append(Paragraph(f"<b>{idx}. {vocab.word}</b>", normal_style))
            elements.append(Paragraph(f"<b>Meaning:</b> {vocab.meaning}", normal_style))
            
            if vocab.synonym:
                elements.append(Paragraph(f"<b>Synonym:</b> {vocab.synonym}", normal_style))
            
            if vocab.sentences:
                elements.append(Paragraph(f"<b>Example Sentences:</b> {vocab.sentences}", normal_style))
            
            # Metadata
            metadata_parts = []
            if vocab.category:
                metadata_parts.append(f"Category: {vocab.category}")
            if vocab.genre:
                metadata_parts.append(f"Genre: {vocab.genre}")
            if vocab.tags:
                metadata_parts.append(f"Tags: {', '.join(vocab.tags)}")
            
            if metadata_parts:
                elements.append(Paragraph(f"<i>{' | '.join(metadata_parts)}</i>", normal_style))
            
            elements.append(Spacer(1, 0.15*inch))
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()

